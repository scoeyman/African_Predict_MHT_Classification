# Import Libraries
import openpyxl
import statistics

# Function for filling missing categoric values
def fill_missing_and_copy(source_sheet, target_sheet, fill_strategy):
    """
    Copies values from the excel source sheet (sheet 1) to the target sheet (sheets 2, 3, 4) and fills missing values.
    
    source_sheet: Source worksheet
    target_sheet: Target worksheet
    fill_strategy: Function to compute the fill value (e.g., mode, median, mean)
    """
    for col_idx, col in enumerate(source_sheet.iter_cols(values_only=True), start=1):
        # Filter numeric values only
        numeric_values = [val for val in col if isinstance(val, (int, float))]
        
        if numeric_values:
            # Compute the fill value based on the strategy
            fill_value = fill_strategy(numeric_values)
        else:
            fill_value = None  # No numeric data to compute
        
        for row_idx, val in enumerate(col, start=1):
            if val is None:
                # Fill missing value
                target_sheet.cell(row=row_idx, column=col_idx, value=fill_value)
            else:
                # Copy original value
                target_sheet.cell(row=row_idx, column=col_idx, value=val)

# Specify the path to your Excel file
file_path = r"C:\Users\scoeyman\Desktop\af predict manual cat imputing\last time (papers)\cleaned_data_with_imputation.xlsx" # Replace this with your actual file path

# Open the workbook
workbook = openpyxl.load_workbook(file_path)

# Get sheets
sheet1 = workbook["Sheet1"]
sheet2 = workbook["Sheet2"]
sheet3 = workbook["Sheet3"]
sheet4 = workbook["Sheet4"]

# Fill Sheet2 using mode
fill_missing_and_copy(sheet1, sheet2, lambda values: statistics.mode(values))

# Fill Sheet3 using median
fill_missing_and_copy(sheet1, sheet3, lambda values: statistics.median(values))

# Fill Sheet4 using mean
fill_missing_and_copy(sheet1, sheet4, lambda values: statistics.mean(values))

# Save the workbook
new_file_path = file_path.replace(".xlsx", "_filled.xlsx")
workbook.save(new_file_path)

print(f"Data processing completed successfully! New file saved as '{new_file_path}'.")


